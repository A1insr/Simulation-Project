import base
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
from scipy.stats import t
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


original_param = {
    'Preoperative Capacity': 30,
    'Emergency Capacity': 10,
    'Emergency Queue Capacity': 10,
    'Laboratory Capacity': 3,
    'Operation Capacity': 50,
    'General Ward Capacity': 65,
    'ICU Capacity': 10,
    'CCU Capacity': 10,
    'Normal Arrival Exp Param': 1,
    'Urgent Arrival Exp Param': 4,
    'Normal Laboratory Param': 1,
    'Urgent Laboratory Param': (10 / 60),
    'After Laboratory Uni a Param': (28 / 60),
    'After Laboratory Uni b Param': (32 / 60),
    'Normal Operation Param': 24,
    'Urgent Operation trgl LB Param': (5 / 60),
    'Urgent Operation trgl M Param': (75 / 60),
    'Urgent Operation trgl UB Param': (100 / 60),
    'Simple Operation Mean': 30.22,
    'Simple Operation SD': 4.96,
    'Medium Operation Mean': 74.54,
    'Medium Operation SD': 9.53,
    'Complex Operation Mean': 242.03,
    'Complex Operation SD': 63.27,
    'Care Unit Exp Param': (1 / 25),
    'End of Service Exp Param': (1 / 40)
}


def run_simulation(simulation_time, param, excel_creation=False):

    """
    This function runs a single hospital simulation and prints detailed metrics about its performance.

    Parameters:
        simulation_time (int): The total duration for which the simulation runs.
        param (dict): A set of parameters used to configure the simulation (e.g., hospital settings, patient flows, etc.).

    Key Steps:
        1. Calls the base simulation function with the provided parameters, enabling Excel creation.
        2. Prints various metrics related to the simulation, including:
           - System-wide metrics: Average time in the system, probabilities, and percentages of patient outcomes.
           - Department productivity: Average productivity (rho) of various hospital departments.
           - Queue performance:
               - Average queue lengths.
               - Maximum queue lengths.
               - Average waiting times in queues.
               - Maximum waiting times in queues.
    """
    simulation = base.simulation(simulation_time, param, excel_creation)['Results']

    print('--------------------------------------------------------------')
    print('Metrics:\n')
    print(f"The average time in the system is: {simulation['average_time_in_system']}")
    print(f"The possibility that the emergency queue capacity is full is: {simulation['Full_Emergency_Queue_Probability']}")
    print(f"The average number of reoperations for patients with complex operations is: {simulation['average_complex_operation_reoperations']}")
    print(f"The percentage of emergency patients who are admitted immediately is: {simulation['immediately_admitted_emergency_patients_percentage']}")

    print('\nAverage productivity of different hospital departments:')
    print(f"rho_Emergency = {simulation['rho_Emergency']}")
    print(f"rho_Preoperative = {simulation['rho_Preoperative']}")
    print(f"rho_Laboratory = {simulation['rho_Laboratory']}")
    print(f"rho_Operation = {simulation['rho_Operation']}")
    print(f"rho_General_Ward = {simulation['rho_General_Ward']}")
    print(f"rho_ICU = {simulation['rho_ICU']}")
    print(f"rho_CCU = {simulation['rho_CCU']}")

    print('\nAverage queue length in different hospital departments:')
    print(f"Lq_Emergency = {simulation['Lq_Emergency']}")
    print(f"Lq_Preoperative = {simulation['Lq_Preoperative']}")
    print(f"Lq_Laboratory_Normal = {simulation['Lq_Laboratory_Normal']}")
    print(f"Lq_Laboratory_Urgent = {simulation['Lq_Laboratory_Urgent']}")
    print(f"Lq_Operation_Normal = {simulation['Lq_Operation_Normal']}")
    print(f"Lq_Operation_Urgent = {simulation['Lq_Operation_Urgent']}")
    print(f"Lq_General_Ward = {simulation['Lq_General_Ward']}")
    print(f"Lq_ICU = {simulation['Lq_ICU']}")
    print(f"Lq_CCU = {simulation['Lq_CCU']}")

    print('\nMaximum queue length in different hospital departments:')
    print(f"Max_Lq_Emergency = {simulation['Max_Lq_Emergency']}")
    print(f"Max_Lq_Preoperative = {simulation['Max_Lq_Preoperative']}")
    print(f"Max_Lq_Laboratory_Normal = {simulation['Max_Lq_Laboratory_Normal']}")
    print(f"Max_Lq_Laboratory_Urgent = {simulation['Max_Lq_Laboratory_Urgent']}")
    print(f"Max_Lq_Operation_Normal = {simulation['Max_Lq_Operation_Normal']}")
    print(f"Max_Lq_Operation_Urgent = {simulation['Max_Lq_Operation_Urgent']}")
    print(f"Max_Lq_General_Ward = {simulation['Max_Lq_General_Ward']}")
    print(f"Max_Lq_ICU = {simulation['Max_Lq_ICU']}")
    print(f"Max_Lq_CCU = {simulation['Max_Lq_CCU']}")

    print('\nAverage waiting time in queues in different hospital departments:')
    print(f"Wq_Emergency = {simulation['Wq_Emergency']}")
    print(f"Wq_Preoperative = {simulation['Wq_Preoperative']}")
    print(f"Wq_Laboratory_Normal = {simulation['Wq_Laboratory_Normal']}")
    print(f"Wq_Laboratory_Urgent = {simulation['Wq_Laboratory_Urgent']}")
    print(f"Wq_Operation_Normal = {simulation['Wq_Operation_Normal']}")
    print(f"Wq_Operation_Urgent = {simulation['Wq_Operation_Urgent']}")
    print(f"Wq_General_Ward = {simulation['Wq_General_Ward']}")
    print(f"Wq_ICU = {simulation['Wq_ICU']}")
    print(f"Wq_CCU = {simulation['Wq_CCU']}")

    print('\nMaximum waiting time in queues in different hospital departments:')
    print(f"Max_Wq_Emergency = {simulation['Max_Wq_Emergency']}")
    print(f"Max_Wq_Preoperative = {simulation['Max_Wq_Preoperative']}")
    print(f"Max_Wq_Laboratory_Normal = {simulation['Max_Wq_Laboratory_Normal']}")
    print(f"Max_Wq_Laboratory_Urgent = {simulation['Max_Wq_Laboratory_Urgent']}")
    print(f"Max_Wq_Operation_Normal = {simulation['Max_Wq_Operation_Normal']}")
    print(f"Max_Wq_Operation_Urgent = {simulation['Max_Wq_Operation_Urgent']}")
    print(f"Max_Wq_General_Ward = {simulation['Max_Wq_General_Ward']}")
    print(f"Max_Wq_ICU = {simulation['Max_Wq_ICU']}")
    print(f"Max_Wq_CCU = {simulation['Max_Wq_CCU']}")
    
    print('\nWarm Period Criteria:')
    print(f"Lq_Preoperative_Warm_Period = {simulation['Lq_Preoperative_Warm_Period']}")
    print(f"Wq_Preoperative_Warm_Period = {simulation['Wq_Preoperative_Warm_Period']}")
    print(f"Finished_Patients = {simulation['Finished_Patients']}")

    print('--------------------------------------------------------------')
    print('Simulation Ended!')


def replication(simulation_time, r, param, alpha):
    """
    Performs multiple replications of the hospital simulation to assess variability and provide confidence intervals for key metrics.

    Parameters:
        simulation_time (int): Duration of each simulation run.
        r (int): The number of independent replications to perform.
        param (dict): Parameters used for the simulation.
        alpha (float): The significance level (e.g., 0.05 for a 95% confidence interval).

    Key Steps:
        1. Run multiple replications of the simulation, storing results for each metric.
        2. Organize the results into a Pandas DataFrame where rows represent metrics and columns represent replications.
        3. Compute statistical metrics:
            - Point estimates (means) and standard deviations.
            - Confidence intervals using the t-distribution.
        4. Add the calculated point estimates and confidence intervals to the DataFrame.

    Returns:
        pd.DataFrame: A DataFrame containing:
            - Metric values for each replication.
            - Point estimates (means).
            - Confidence intervals for each metric.
    """

    # Initialize results dictionary
    list_of_result = None

    for i in tqdm(range(r)):
        # Run the simulation
        result = base.simulation(simulation_time, param)['Results']

        # On the first iteration, initialize structures
        if i == 0:
            list_of_result = {key: [0] * r for key in result.keys()}

        # Store the result for the current replication
        for key in result.keys():
            list_of_result[key][i] = result[key]

    # Create a DataFrame where rows correspond to metrics and columns to replications
    results = pd.DataFrame(list_of_result).transpose()
    results.columns = [f"Replication{j + 1}" for j in range(r)]

    # Calculate point estimate (mean) and confidence intervals for each metric
    means = results.mean(axis=1)  # Point estimate
    stds = results.std(axis=1)  # Standard deviation
    n = r  # Number of replications
    t_alpha = t.ppf(1 - alpha / 2, df=n - 1)  # Critical value from t-distribution
    ci_half_width = t_alpha * stds / (n ** 0.5)  # Half-width of the confidence interval

    # Add the point estimate and CI to the DataFrame
    results['Point Estimate'] = means
    results['Confidence Interval'] = [
        f"[{round(mean - ci, 4)}, {round(mean + ci, 4)}]"
        for mean, ci in zip(means, ci_half_width)
    ]

    # Save the results DataFrame as an Excel file (including row names)
    file_name = "simulation_results.xlsx"
    results.to_excel(file_name, sheet_name="Replication Results", index=True)
    print(f"Results saved to {file_name}")

    # Load the Excel file to apply formatting using openpyxl
    wb = load_workbook(file_name)
    sheet = wb["Replication Results"]

    # Change font to Times New Roman and center align all cells
    font = Font(name='Times New Roman', size=12)
    alignment = Alignment(horizontal='center', vertical='center')

    # Apply to all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # Adjust column widths based on the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Adding a bit of padding
        sheet.column_dimensions[column].width = adjusted_width

    # Apply color to the last two columns (Point Estimate and Confidence Interval)
    fill_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    last_column = len(results.columns) + 1  # Account for the index being included
    second_last_column = len(results.columns)  # Adjusted for index

    # Color the last two columns
    for row in sheet.iter_rows(min_col=second_last_column, max_col=last_column):
        for cell in row:
            cell.fill = fill_color

    # Save the modified workbook with adjusted column sizes and colored cells
    wb.save(file_name)
    print(f"Excel file formatted, column widths adjusted, and saved as {file_name}")

    return results


def multi_sensitivity_analysis_with_individual_plots(simulation_time, param, analyses, replications, alpha=0.05):
    """
    Perform multiple sensitivity analyses and save individual plots for each metric/parameter pair.

    Args:
        simulation_time: Total simulation time (e.g., 30 * 24 for 30 days).
        param: Original parameter dictionary.
        analyses: List of dictionaries, where each dict has:
                  - 'metric': Metric to analyse (e.g., 'average_time_in_system')
                  - 'parameter_name': Name of the parameter to vary (e.g., 'Emergency Queue Capacity')
                  - 'parameter_values': List of values for the parameter
        replications: Number of replications per parameter value.
        alpha: Significance level for confidence intervals.

    Returns:
        A list of DataFrames, one for each analysis, containing the results.
    """
    results_list = []

    for analysis in analyses:
        parameter_name = analysis['parameter_name']
        parameter_values = analysis['parameter_values']
        metric = analysis['metric']
        results = []

        for value in tqdm(parameter_values, desc=f"Analyzing {metric} ({parameter_name})"):
            param_copy = param.copy()
            param_copy[parameter_name] = value
            metrics = []

            for _ in range(replications):
                result = base.simulation(simulation_time, param_copy)['Results']
                metrics.append(result[metric])

            # Calculate statistics
            mean_metric = np.mean(metrics)
            std_metric = np.std(metrics, ddof=1)
            t_alpha = t.ppf(1 - alpha / 2, df=replications - 1)
            ci_half_width = t_alpha * (std_metric / np.sqrt(replications))

            results.append({
                'Parameter Value': value,
                'Point Estimate': mean_metric,
                'Lower CI': mean_metric - ci_half_width,
                'Upper CI': mean_metric + ci_half_width
            })

        df_results = pd.DataFrame(results)
        results_list.append(df_results)

        # Plot and save each analysis
        plt.figure(figsize=(8, 6))
        plt.plot(df_results['Parameter Value'], df_results['Point Estimate'], label='Point Estimate', color='orange',
                 linestyle='--')
        plt.fill_between(
            df_results['Parameter Value'],
            df_results['Lower CI'],
            df_results['Upper CI'],
            color='blue',
            alpha=0.2,
            label='Confidence Interval'
        )
        plt.title(f"Sensitivity Analysis: {metric} vs {parameter_name}")
        plt.xlabel(parameter_name)
        plt.ylabel(metric)
        plt.legend()
        plt.grid(True)

        # Save each plot as a separate file
        file_name = f"sensitivity_{metric}_vs_{parameter_name}.svg"
        plt.savefig(file_name, format="svg", bbox_inches="tight")
        print(f"Saved plot as {file_name}")
        plt.show()
        plt.close()  # Close the figure to free memory

    return results_list


# Run multi-sensitivity analysis and save individual plots.
analyses = [
    {
        'metric': 'average_time_in_system',
        'parameter_name': 'Normal Arrival Exp Param',
        'parameter_values': np.linspace(0.1, 7, 15)
    },
    {
        'metric': 'Lq_Preoperative',
        'parameter_name': 'Normal Arrival Exp Param',
        'parameter_values': np.linspace(0.1, 7, 15)
    },
    {
        'metric': 'immediately_admitted_emergency_patients_percentage',
        'parameter_name': 'Emergency Capacity',
        'parameter_values': np.arange(1, 4, 1)
    },
    {
        'metric': 'rho_Emergency',
        'parameter_name': 'Emergency Capacity',
        'parameter_values': np.arange(1, 11, 1)
    },
    {
        'metric': 'rho_Laboratory',
        'parameter_name': 'Normal Laboratory Param',
        'parameter_values': np.linspace(0.5, 5, 10)
    }
]

if __name__ == "__main__":

    results = multi_sensitivity_analysis_with_individual_plots(
        simulation_time=30 * 24,
        param=original_param,
        analyses=analyses,
        replications=10
    )

    # run simulation once, show metrics results and save the trace as an Excel file (output.xlsx).
    run_simulation((30 * 24), original_param, excel_creation=True)

    # run some replication of the model for obtaining point estimate and confidence interval estimate for each metric.
    # save those as an Excel file (simulation_results.xlsx).
    result = replication((30 * 24), 25, original_param, 0.05)
